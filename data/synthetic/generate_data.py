"""
Pharma DR · Synthetic Data Generator
=====================================
Generates realistic pharmaceutical sales data for República Dominicana.
Produces ~100,000 fact_sales rows spanning 2021-2024 + 200 clients.

Run: python data/synthetic/generate_data.py
Output: data/synthetic/clients.csv, sales.csv, commissions.csv
        + 6 Excel distributor format samples
"""

import os
import sys
import random
import json
from datetime import date, timedelta, datetime
from pathlib import Path
from typing import List, Dict, Any

import numpy as np
import pandas as pd
from faker import Faker

# ── Project root on path ──────────────────────────────────────
ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

fake = Faker("es_MX")   # Spanish names (close enough for DR)
Faker.seed(42)
random.seed(42)
np.random.seed(42)

OUTPUT_DIR    = ROOT / "data" / "synthetic"
EXCEL_DIR     = ROOT / "data" / "excel_samples"
MAPPING_DIR   = ROOT / "data" / "mappings"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_DIR.mkdir(parents=True, exist_ok=True)
MAPPING_DIR.mkdir(parents=True, exist_ok=True)

# ════════════════════════════════════════════════════════════════
# MASTER REFERENCE DATA
# ════════════════════════════════════════════════════════════════

ZONES = {
    "CAP": {"name": "Capital",  "weight": 0.40},
    "NOR": {"name": "Norte",    "weight": 0.25},
    "EST": {"name": "Este",     "weight": 0.15},
    "SUR": {"name": "Sur",      "weight": 0.12},
    "OES": {"name": "Oeste",    "weight": 0.08},
}

CITIES = [
    {"name": "Santo Domingo",         "province": "Distrito Nacional",     "zone": "CAP", "weight": 0.22},
    {"name": "Santo Domingo Este",    "province": "Santo Domingo",         "zone": "CAP", "weight": 0.10},
    {"name": "Santo Domingo Norte",   "province": "Santo Domingo",         "zone": "CAP", "weight": 0.06},
    {"name": "Santo Domingo Oeste",   "province": "Santo Domingo",         "zone": "CAP", "weight": 0.04},
    {"name": "Boca Chica",            "province": "Santo Domingo",         "zone": "CAP", "weight": 0.02},
    {"name": "Santiago",              "province": "Santiago",              "zone": "NOR", "weight": 0.14},
    {"name": "La Vega",               "province": "La Vega",               "zone": "NOR", "weight": 0.04},
    {"name": "Puerto Plata",          "province": "Puerto Plata",          "zone": "NOR", "weight": 0.03},
    {"name": "San Francisco de Macorís","province": "Duarte",              "zone": "NOR", "weight": 0.03},
    {"name": "Moca",                  "province": "Espaillat",             "zone": "NOR", "weight": 0.02},
    {"name": "La Romana",             "province": "La Romana",             "zone": "EST", "weight": 0.06},
    {"name": "San Pedro de Macorís",  "province": "San Pedro de Macorís",  "zone": "EST", "weight": 0.05},
    {"name": "Higüey",                "province": "La Altagracia",         "zone": "EST", "weight": 0.03},
    {"name": "Hato Mayor",            "province": "Hato Mayor",            "zone": "EST", "weight": 0.01},
    {"name": "Barahona",              "province": "Barahona",              "zone": "SUR", "weight": 0.04},
    {"name": "San Cristóbal",         "province": "San Cristóbal",         "zone": "SUR", "weight": 0.04},
    {"name": "Azua",                  "province": "Azua",                  "zone": "SUR", "weight": 0.02},
    {"name": "Bani",                  "province": "Peravia",               "zone": "SUR", "weight": 0.02},
    {"name": "Monte Cristi",          "province": "Monte Cristi",          "zone": "OES", "weight": 0.03},
    {"name": "Mao",                   "province": "Valverde",              "zone": "OES", "weight": 0.02},
    {"name": "Dajabón",               "province": "Dajabón",               "zone": "OES", "weight": 0.01},
    {"name": "Nagua",                 "province": "María Trinidad Sánchez","zone": "NOR", "weight": 0.02},
    {"name": "Bonao",                 "province": "Monseñor Nouel",        "zone": "NOR", "weight": 0.02},
    {"name": "Cotui",                 "province": "Sánchez Ramírez",       "zone": "NOR", "weight": 0.01},
]
CITY_WEIGHTS = [c["weight"] for c in CITIES]

PRODUCTS = [
    {"id": "SAP-001", "name": "Amoxicilina 500mg Cápsulas",    "cat": "ANTIBIOTICO",    "lab": "GENFAR",  "rx": "RX",  "cost": 28.50,  "price": 58.00},
    {"id": "SAP-002", "name": "Ciprofloxacino 500mg Tabletas", "cat": "ANTIBIOTICO",    "lab": "BAYER",   "rx": "RX",  "cost": 95.00,  "price": 195.00},
    {"id": "SAP-003", "name": "Azitromicina 500mg Tabletas",   "cat": "ANTIBIOTICO",    "lab": "PFIZER",  "rx": "RX",  "cost": 145.00, "price": 295.00},
    {"id": "SAP-004", "name": "Clindamicina 300mg Cápsulas",   "cat": "ANTIBIOTICO",    "lab": "PFIZER",  "rx": "RX",  "cost": 82.00,  "price": 168.00},
    {"id": "SAP-005", "name": "Ceftriaxona 1g Inyectable",     "cat": "ANTIBIOTICO",    "lab": "ROCHE",   "rx": "RX",  "cost": 185.00, "price": 380.00},
    {"id": "SAP-006", "name": "Metronidazol 500mg Tabletas",   "cat": "ANTIBIOTICO",    "lab": "SANOFI",  "rx": "RX",  "cost": 22.00,  "price": 45.00},
    {"id": "SAP-007", "name": "Ibuprofeno 400mg Tabletas",     "cat": "ANALGESICO",     "lab": "ABBOTT",  "rx": "OTC", "cost": 18.00,  "price": 38.00},
    {"id": "SAP-008", "name": "Paracetamol 500mg Tabletas",    "cat": "ANALGESICO",     "lab": "GSK",     "rx": "OTC", "cost": 8.50,   "price": 18.00},
    {"id": "SAP-009", "name": "Diclofenaco 50mg Tabletas",     "cat": "ANALGESICO",     "lab": "NOVARTIS","rx": "OTC", "cost": 25.00,  "price": 52.00},
    {"id": "SAP-010", "name": "Naproxeno 500mg Tabletas",      "cat": "ANALGESICO",     "lab": "ROCHE",   "rx": "OTC", "cost": 32.00,  "price": 65.00},
    {"id": "SAP-011", "name": "Tramadol 50mg Cápsulas",        "cat": "ANALGESICO",     "lab": "MERCK",   "rx": "RX",  "cost": 68.00,  "price": 140.00},
    {"id": "SAP-012", "name": "Vitamina C 500mg Tabletas",     "cat": "VITAMINAS",      "lab": "BAYER",   "rx": "OTC", "cost": 35.00,  "price": 72.00},
    {"id": "SAP-013", "name": "Vitamina D3 1000UI Cápsulas",   "cat": "VITAMINAS",      "lab": "ABBOTT",  "rx": "OTC", "cost": 42.00,  "price": 88.00},
    {"id": "SAP-014", "name": "Complejo B Tabletas",           "cat": "VITAMINAS",      "lab": "MERCK",   "rx": "OTC", "cost": 38.00,  "price": 78.00},
    {"id": "SAP-015", "name": "Zinc 20mg Tabletas",            "cat": "VITAMINAS",      "lab": "MEDCO",   "rx": "OTC", "cost": 22.00,  "price": 45.00},
    {"id": "SAP-016", "name": "Calcio + D3 600mg Tabletas",    "cat": "VITAMINAS",      "lab": "ABBOTT",  "rx": "OTC", "cost": 65.00,  "price": 135.00},
    {"id": "SAP-017", "name": "Multivitamínico Adultos",       "cat": "VITAMINAS",      "lab": "PFIZER",  "rx": "OTC", "cost": 85.00,  "price": 175.00},
    {"id": "SAP-018", "name": "Atorvastatina 20mg Tabletas",   "cat": "CARDIOVASCULAR", "lab": "PFIZER",  "rx": "RX",  "cost": 58.00,  "price": 120.00},
    {"id": "SAP-019", "name": "Losartán 50mg Tabletas",        "cat": "CARDIOVASCULAR", "lab": "MERCK",   "rx": "RX",  "cost": 45.00,  "price": 92.00},
    {"id": "SAP-020", "name": "Metoprolol 50mg Tabletas",      "cat": "CARDIOVASCULAR", "lab": "ASTRA",   "rx": "RX",  "cost": 38.00,  "price": 78.00},
    {"id": "SAP-021", "name": "Amlodipino 10mg Tabletas",      "cat": "CARDIOVASCULAR", "lab": "PFIZER",  "rx": "RX",  "cost": 48.00,  "price": 98.00},
    {"id": "SAP-022", "name": "Enalapril 10mg Tabletas",       "cat": "CARDIOVASCULAR", "lab": "MERCK",   "rx": "RX",  "cost": 32.00,  "price": 65.00},
    {"id": "SAP-023", "name": "Clopidogrel 75mg Tabletas",     "cat": "CARDIOVASCULAR", "lab": "SANOFI",  "rx": "RX",  "cost": 95.00,  "price": 195.00},
    {"id": "SAP-024", "name": "Furosemida 40mg Tabletas",      "cat": "CARDIOVASCULAR", "lab": "SANOFI",  "rx": "RX",  "cost": 15.00,  "price": 32.00},
    {"id": "SAP-025", "name": "Metformina 850mg Tabletas",     "cat": "DIABETES",       "lab": "MERCK",   "rx": "RX",  "cost": 28.00,  "price": 58.00},
    {"id": "SAP-026", "name": "Glibenclamida 5mg Tabletas",    "cat": "DIABETES",       "lab": "ROCHE",   "rx": "RX",  "cost": 18.00,  "price": 38.00},
    {"id": "SAP-027", "name": "Insulina NPH 100UI/ml",         "cat": "DIABETES",       "lab": "LILLY",   "rx": "RX",  "cost": 285.00, "price": 585.00},
    {"id": "SAP-028", "name": "Sitagliptina 100mg Tabletas",   "cat": "DIABETES",       "lab": "MERCK",   "rx": "RX",  "cost": 485.00, "price": 995.00},
    {"id": "SAP-029", "name": "Salbutamol 100mcg Inhalador",   "cat": "RESPIRATORIO",   "lab": "GSK",     "rx": "RX",  "cost": 195.00, "price": 400.00},
    {"id": "SAP-030", "name": "Loratadina 10mg Tabletas",      "cat": "RESPIRATORIO",   "lab": "MERCK",   "rx": "OTC", "cost": 32.00,  "price": 65.00},
    {"id": "SAP-031", "name": "Montelukast 10mg Tabletas",     "cat": "RESPIRATORIO",   "lab": "MERCK",   "rx": "RX",  "cost": 245.00, "price": 500.00},
    {"id": "SAP-032", "name": "Fluticasona 50mcg Spray Nasal", "cat": "RESPIRATORIO",   "lab": "GSK",     "rx": "RX",  "cost": 385.00, "price": 790.00},
    {"id": "SAP-033", "name": "Ambroxol 30mg Jarabe",          "cat": "RESPIRATORIO",   "lab": "MEDCO",   "rx": "OTC", "cost": 145.00, "price": 298.00},
    {"id": "SAP-034", "name": "Omeprazol 20mg Cápsulas",       "cat": "GASTRO",         "lab": "ASTRA",   "rx": "OTC", "cost": 42.00,  "price": 88.00},
    {"id": "SAP-035", "name": "Ranitidina 150mg Tabletas",     "cat": "GASTRO",         "lab": "GSK",     "rx": "OTC", "cost": 28.00,  "price": 58.00},
    {"id": "SAP-036", "name": "Metoclopramida 10mg Tabletas",  "cat": "GASTRO",         "lab": "SANOFI",  "rx": "OTC", "cost": 18.00,  "price": 38.00},
    {"id": "SAP-037", "name": "Loperamida 2mg Cápsulas",       "cat": "GASTRO",         "lab": "JNJ",     "rx": "OTC", "cost": 48.00,  "price": 98.00},
    {"id": "SAP-038", "name": "Pantoprazol 40mg Tabletas",     "cat": "GASTRO",         "lab": "PFIZER",  "rx": "RX",  "cost": 95.00,  "price": 195.00},
    {"id": "SAP-039", "name": "Hidrocortisona 1% Crema",       "cat": "DERMATOLOGIA",   "lab": "JNJ",     "rx": "OTC", "cost": 55.00,  "price": 115.00},
    {"id": "SAP-040", "name": "Clotrimazol 1% Crema",          "cat": "DERMATOLOGIA",   "lab": "BAYER",   "rx": "OTC", "cost": 48.00,  "price": 98.00},
    {"id": "SAP-041", "name": "Aciclovir 5% Crema",            "cat": "DERMATOLOGIA",   "lab": "GSK",     "rx": "OTC", "cost": 145.00, "price": 298.00},
    {"id": "SAP-042", "name": "Betametasona 0.05% Crema",      "cat": "DERMATOLOGIA",   "lab": "MERCK",   "rx": "RX",  "cost": 75.00,  "price": 155.00},
    {"id": "SAP-043", "name": "Albendazol 400mg Tabletas",     "cat": "ANTIPARASITARIO","lab": "GSK",     "rx": "OTC", "cost": 38.00,  "price": 78.00},
    {"id": "SAP-044", "name": "Ivermectina 6mg Tabletas",      "cat": "ANTIPARASITARIO","lab": "LAFRANCOL","rx": "RX", "cost": 55.00,  "price": 115.00},
    {"id": "SAP-045", "name": "Dexametasona 4mg Inyectable",   "cat": "CORTICOIDE",     "lab": "MERCK",   "rx": "RX",  "cost": 18.00,  "price": 38.00},
    {"id": "SAP-046", "name": "Prednisona 20mg Tabletas",      "cat": "CORTICOIDE",     "lab": "MERCK",   "rx": "RX",  "cost": 25.00,  "price": 52.00},
    {"id": "SAP-047", "name": "Levotiroxina 100mcg Tabletas",  "cat": "HORMONA",        "lab": "ABBOTT",  "rx": "RX",  "cost": 185.00, "price": 380.00},
    {"id": "SAP-048", "name": "Alprazolam 0.5mg Tabletas",     "cat": "PSIQUIATRIA",    "lab": "PFIZER",  "rx": "RX",  "cost": 68.00,  "price": 140.00},
    {"id": "SAP-049", "name": "Fluoxetina 20mg Cápsulas",      "cat": "PSIQUIATRIA",    "lab": "LILLY",   "rx": "RX",  "cost": 78.00,  "price": 160.00},
    {"id": "SAP-050", "name": "Amitriptilina 25mg Tabletas",   "cat": "PSIQUIATRIA",    "lab": "MEDCO",   "rx": "RX",  "cost": 22.00,  "price": 45.00},
]

# Product weights (more common products sell more)
PRODUCT_WEIGHTS = [
    6, 4, 3, 2, 2, 3,      # ANTIBIOTICO
    8, 10, 6, 4, 2,         # ANALGESICO
    7, 5, 6, 4, 4, 5,       # VITAMINAS
    5, 4, 3, 3, 3, 2, 3,   # CARDIOVASCULAR
    5, 3, 2, 2,             # DIABETES
    4, 5, 3, 2, 3,          # RESPIRATORIO
    6, 4, 4, 3, 4,          # GASTRO
    3, 4, 2, 2,             # DERMATOLOGIA
    3, 2,                   # ANTIPARASITARIO
    3, 2, 2, 1, 2, 2        # CORTICOIDE / HORMONA / PSIQ
]

SALESPERSONS = [
    {"id": "SP-001", "name": "Carlos Batista Pérez",     "zone": "CAP", "target": 850000,  "rate": 0.040},
    {"id": "SP-002", "name": "María González Sánchez",   "zone": "CAP", "target": 820000,  "rate": 0.038},
    {"id": "SP-003", "name": "José Rodríguez Mejía",     "zone": "CAP", "target": 900000,  "rate": 0.042},
    {"id": "SP-004", "name": "Ana Martínez Ortiz",       "zone": "CAP", "target": 750000,  "rate": 0.035},
    {"id": "SP-005", "name": "Luis Fernández Taveras",   "zone": "NOR", "target": 800000,  "rate": 0.040},
    {"id": "SP-006", "name": "Carmen Valdez Peralta",    "zone": "NOR", "target": 780000,  "rate": 0.038},
    {"id": "SP-007", "name": "Roberto Jiménez Cruz",     "zone": "NOR", "target": 950000,  "rate": 0.045},
    {"id": "SP-008", "name": "Patricia Almonte Rivas",   "zone": "NOR", "target": 720000,  "rate": 0.035},
    {"id": "SP-009", "name": "Miguel Ángel Rosario",     "zone": "EST", "target": 820000,  "rate": 0.040},
    {"id": "SP-010", "name": "Sandra Corporán Díaz",     "zone": "EST", "target": 780000,  "rate": 0.038},
    {"id": "SP-011", "name": "Ramón Féliz Cepeda",       "zone": "EST", "target": 760000,  "rate": 0.036},
    {"id": "SP-012", "name": "Yolanda Tejeda Núñez",     "zone": "SUR", "target": 760000,  "rate": 0.038},
    {"id": "SP-013", "name": "Eduardo Pichardo Reyes",   "zone": "SUR", "target": 720000,  "rate": 0.035},
    {"id": "SP-014", "name": "Cecilia Marte Herrera",    "zone": "SUR", "target": 700000,  "rate": 0.033},
    {"id": "SP-015", "name": "Francisco Castillo López", "zone": "OES", "target": 800000,  "rate": 0.040},
]

# Zones → salespersons mapping
ZONE_SP = {}
for sp in SALESPERSONS:
    ZONE_SP.setdefault(sp["zone"], []).append(sp["id"])

DISTRIBUTORS = [
    {"code": "INT",    "name": "Ventas Internas",         "type": "INTERNO"},
    {"code": "DIST_A", "name": "Distribuidora Ramos",     "type": "EXTERNO"},
    {"code": "DIST_B", "name": "MediFar Dominicana",      "type": "EXTERNO"},
    {"code": "DIST_C", "name": "Farmacorp",               "type": "EXTERNO"},
    {"code": "DIST_D", "name": "AlphaFarma Group",        "type": "EXTERNO"},
    {"code": "DIST_E", "name": "BioPharma Distribution",  "type": "EXTERNO"},
    {"code": "DIST_F", "name": "MedDist Nacional",        "type": "EXTERNO"},
]
DIST_WEIGHTS = [0.50, 0.10, 0.10, 0.08, 0.08, 0.07, 0.07]

CLIENT_TYPES = [
    ("FARMACIA_INDEPENDIENTE", 0.45),
    ("CADENA_FARMACIA",        0.25),
    ("HOSPITAL_PUBLICO",       0.10),
    ("CLINICA_PRIVADA",        0.12),
    ("DISTRIBUIDORA_LOCAL",    0.08),
]

DR_SURNAMES = [
    "García","Rodríguez","Martínez","Hernández","López",
    "González","Pérez","Sánchez","Ramírez","Torres",
    "Flores","Díaz","Cruz","Reyes","Morales",
    "Jiménez","Medina","Cabrera","Taveras","Almonte",
    "Batista","Marte","Féliz","Pichardo","Corporán",
    "Valdez","Fernández","Ortiz","Tejeda","Polanco",
    "Castillo","Méndez","Cepeda","Peralta","Herrera",
]

# ════════════════════════════════════════════════════════════════
# GENERATOR FUNCTIONS
# ════════════════════════════════════════════════════════════════

def generate_rnc() -> str:
    """Generate a fake Dominican RNC (tax ID)."""
    return f"1-{random.randint(10,99):02d}-{random.randint(10000,99999):05d}-{random.randint(0,9)}"


def generate_clients(n: int = 200) -> pd.DataFrame:
    """Generate n pharmacy/hospital clients across DR cities."""
    records = []
    city_weights_norm = np.array(CITY_WEIGHTS) / sum(CITY_WEIGHTS)

    for i in range(1, n + 1):
        city = np.random.choice(CITIES, p=city_weights_norm)
        ctype, _ = random.choices(CLIENT_TYPES, weights=[w for _, w in CLIENT_TYPES])[0], None
        ctype = random.choices([ct for ct, _ in CLIENT_TYPES],
                               weights=[w for _, w in CLIENT_TYPES])[0]

        if ctype == "FARMACIA_INDEPENDIENTE":
            suffix = random.choice(["Farmacia", "Farmácia", "Droguería"])
            surname = random.choice(DR_SURNAMES)
            name = f"{suffix} {surname}"
        elif ctype == "CADENA_FARMACIA":
            name = random.choice([
                "Carol Ana Farmacia", "Farmacias Estrella", "Farmacia Cruz Verde",
                "Grupo Fármacos", "Farmacias del Pueblo", "Farmacia Nacional",
                "Fármaco Express", "Clínica Farmacéutica", "Farmacia Royal",
                "Droguería Central",
            ]) + f" #{i}"
        elif ctype == "HOSPITAL_PUBLICO":
            name = random.choice([
                "Hospital General", "Hospital Regional", "Centro de Salud",
                "Hospital Universitario", "Hospital Materno"
            ]) + f" {city['city_name']}"
        elif ctype == "CLINICA_PRIVADA":
            name = f"Clínica {random.choice(DR_SURNAMES)} & Asociados"
        else:
            name = f"Distribuidora Local {random.choice(DR_SURNAMES)}"

        credit_limit = {
            "FARMACIA_INDEPENDIENTE": random.uniform(30000, 150000),
            "CADENA_FARMACIA":        random.uniform(200000, 800000),
            "HOSPITAL_PUBLICO":       random.uniform(500000, 2000000),
            "CLINICA_PRIVADA":        random.uniform(100000, 400000),
            "DISTRIBUIDORA_LOCAL":    random.uniform(80000, 300000),
        }[ctype]

        records.append({
            "client_id":    f"CLI-{i:04d}",
            "client_name":  name,
            "client_type":  ctype,
            "rnc":          generate_rnc(),
            "address":      f"{fake.street_address()}, {city['city_name']}",
            "phone":        f"809-{random.randint(200,799):03d}-{random.randint(1000,9999):04d}",
            "email":        f"compras@{name.lower().replace(' ','').replace('#','')[:15]}.com",
            "credit_limit": round(credit_limit, 2),
            "payment_terms": random.choice([15, 30, 45, 60]),
            "city_name":    city["name"],
            "province":     city["province"],
            "zone_code":    city["zone"],
        })

    return pd.DataFrame(records)


def pick_seasonal_weight(sale_date: date) -> float:
    """
    Pharmaceutical sales in DR are seasonal:
    - Peak: Jan (flu), Apr-May (allergies), Oct-Nov (flu again), Dec (holidays)
    - Low: Jun-Aug (summer vacation)
    """
    month = sale_date.month
    seasonal = {
        1: 1.25, 2: 1.05, 3: 1.00, 4: 1.15, 5: 1.10, 6: 0.90,
        7: 0.88, 8: 0.85, 9: 0.95, 10: 1.10, 11: 1.20, 12: 1.30,
    }
    return seasonal.get(month, 1.0)


def generate_sales(
    clients: pd.DataFrame,
    n_invoices: int = 25000,
    start_date: date = date(2021, 1, 1),
    end_date: date = date(2024, 12, 31),
) -> pd.DataFrame:
    """Generate n_invoices invoices (each 1-6 lines) = ~75k-100k rows."""
    records = []
    total_days = (end_date - start_date).days
    prod_weights = np.array(PRODUCT_WEIGHTS) / sum(PRODUCT_WEIGHTS)
    city_weights_norm = np.array(CITY_WEIGHTS) / sum(CITY_WEIGHTS)

    client_list = clients.to_dict("records")

    invoice_num = 100000
    sale_key = 1

    for _ in range(n_invoices):
        invoice_num += 1
        client = random.choice(client_list)

        # Generate a sale date with seasonal weighting
        while True:
            offset = random.randint(0, total_days)
            sale_date = start_date + timedelta(days=offset)
            if random.random() < pick_seasonal_weight(sale_date) / 1.30:
                break

        # City may differ from client registered city (mobile purchasing)
        if random.random() < 0.85:
            city = next(c for c in CITIES if c["name"] == client["city_name"])
        else:
            city = np.random.choice(CITIES, p=city_weights_norm)

        zone_code = city["zone"]
        salesperson_id = random.choice(ZONE_SP.get(zone_code, [SALESPERSONS[0]["id"]]))
        distributor = random.choices(DISTRIBUTORS, weights=DIST_WEIGHTS)[0]
        source_system = distributor["code"] if distributor["type"] == "EXTERNO" else "SAP_HANA"

        n_lines = random.choices([1, 2, 3, 4, 5, 6], weights=[30, 25, 20, 12, 8, 5])[0]
        selected_products = random.choices(PRODUCTS, weights=prod_weights, k=n_lines)

        for line_num, product in enumerate(selected_products, 1):
            quantity = random.choices(
                [1, 2, 3, 5, 10, 20, 50, 100],
                weights=[10, 15, 20, 20, 15, 10, 7, 3]
            )[0]

            # Add ±15% price variation
            unit_price = round(product["price"] * random.uniform(0.90, 1.10), 2)

            # Discount: larger clients get bigger discounts
            if client["client_type"] in ("HOSPITAL_PUBLICO", "CADENA_FARMACIA"):
                discount_pct = round(random.uniform(5, 20), 2)
            elif client["client_type"] == "DISTRIBUIDORA_LOCAL":
                discount_pct = round(random.uniform(10, 25), 2)
            else:
                discount_pct = round(random.uniform(0, 10), 2)

            gross_amount  = round(quantity * unit_price, 2)
            discount_amt  = round(gross_amount * discount_pct / 100, 2)
            net_amount    = round(gross_amount - discount_amt, 2)
            cost_amount   = round(product["cost"] * quantity, 2)
            margin_amount = round(net_amount - cost_amount, 2)
            margin_pct    = round(margin_amount / net_amount * 100, 3) if net_amount > 0 else 0

            records.append({
                "sale_key":         sale_key,
                "date_key":         int(sale_date.strftime("%Y%m%d")),
                "full_date":        sale_date.isoformat(),
                "product_id":       product["id"],
                "product_name":     product["name"],
                "category":         product["cat"],
                "lab_code":         product["lab"],
                "client_id":        client["client_id"],
                "client_name":      client["client_name"],
                "client_type":      client["client_type"],
                "city_name":        city["name"],
                "zone_code":        zone_code,
                "distributor_code": distributor["code"],
                "distributor_name": distributor["name"],
                "distributor_type": distributor["type"],
                "salesperson_id":   salesperson_id,
                "invoice_number":   f"FAC-{invoice_num:08d}",
                "invoice_line":     line_num,
                "quantity":         quantity,
                "unit_price":       unit_price,
                "gross_amount":     gross_amount,
                "discount_pct":     discount_pct,
                "discount_amount":  discount_amt,
                "net_amount":       net_amount,
                "cost_amount":      cost_amount,
                "margin_amount":    margin_amount,
                "margin_pct":       margin_pct,
                "source_system":    source_system,
                "source_record_id": f"{source_system}-{invoice_num:08d}",
            })
            sale_key += 1

    df = pd.DataFrame(records)
    # Apply YoY growth trend (5% per year)
    df["year"] = pd.to_datetime(df["full_date"]).dt.year
    growth = {2021: 1.00, 2022: 1.05, 2023: 1.10, 2024: 1.16}
    for yr, factor in growth.items():
        mask = df["year"] == yr
        df.loc[mask, "net_amount"]    = (df.loc[mask, "net_amount"] * factor).round(2)
        df.loc[mask, "gross_amount"]  = (df.loc[mask, "gross_amount"] * factor).round(2)
        df.loc[mask, "margin_amount"] = (df.loc[mask, "net_amount"] - df.loc[mask, "cost_amount"]).round(2)
        df.loc[mask, "margin_pct"]    = (df.loc[mask, "margin_amount"] / df.loc[mask, "net_amount"].replace(0, np.nan) * 100).round(3)

    df.drop(columns=["year"], inplace=True)
    return df


def generate_targets(year: int, month: int) -> List[Dict]:
    """Monthly targets per salesperson — slight variation per month."""
    records = []
    for sp in SALESPERSONS:
        seasonal = pick_seasonal_weight(date(year, month, 1))
        target = round(sp["target"] * seasonal * random.uniform(0.95, 1.05), 2)
        records.append({
            "year": year,
            "month": month,
            "salesperson_id": sp["id"],
            "salesperson_name": sp["name"],
            "zone_code": sp["zone"],
            "monthly_target": target,
            "commission_rate": sp["rate"],
        })
    return records


def generate_product_mapping() -> pd.DataFrame:
    """
    Simulate distributor product codes → SAP master product codes.
    Each distributor uses different codes for the same products.
    """
    rows = []
    dist_codes = {
        "DIST_A": lambda pid, i: f"RAM-{i:03d}",
        "DIST_B": lambda pid, i: f"MF{i:04d}",
        "DIST_C": lambda pid, i: f"FC-{pid.replace('SAP-','')}",
        "DIST_D": lambda pid, i: f"ALF{i:03d}A",
        "DIST_E": lambda pid, i: f"BP{i:02d}X",
        "DIST_F": lambda pid, i: f"MD-{i:05d}",
    }
    for i, prod in enumerate(PRODUCTS, 1):
        for dist_code, code_fn in dist_codes.items():
            dist_product_code = code_fn(prod["id"], i)
            rows.append({
                "distributor_code":    dist_code,
                "dist_product_code":   dist_product_code,
                "dist_product_desc":   prod["name"] + random.choice(["", " (Genérico)", " Orig.", " Ref."]),
                "sap_product_id":      prod["id"],
                "sap_product_name":    prod["name"],
                "sap_category":        prod["cat"],
                "mapping_confidence":  random.choice([1.0, 1.0, 1.0, 0.95, 0.90]),
                "validated_flag":      True,
                "last_updated":        date(2024, 1, 15).isoformat(),
            })
    return pd.DataFrame(rows)


# ════════════════════════════════════════════════════════════════
# EXCEL DISTRIBUTOR FORMAT SAMPLES
# ════════════════════════════════════════════════════════════════

def generate_excel_dist_a(sales_df: pd.DataFrame) -> None:
    """
    Distributor A — Ramos: Standard row-per-invoice-line format.
    Columns: Fecha | Num_Factura | Cod_Producto | Descripcion | Cantidad | Precio | Total | Cliente | Ciudad
    """
    mask = sales_df["distributor_code"] == "DIST_A"
    sample = sales_df[mask].head(200).copy()
    df = pd.DataFrame({
        "Fecha":         sample["full_date"],
        "Num_Factura":   sample["invoice_number"],
        "Cod_Producto":  sample["product_id"].str.replace("SAP-", "RAM-"),
        "Descripcion":   sample["product_name"],
        "Cantidad":      sample["quantity"],
        "Precio_Unit":   sample["unit_price"],
        "Descuento_Pct": sample["discount_pct"],
        "Total_Neto":    sample["net_amount"],
        "Cliente":       sample["client_name"],
        "Ciudad":        sample["city_name"],
        "Zona":          sample["zone_code"],
    })
    path = EXCEL_DIR / "DIST_A_Ramos_ventas.xlsx"
    df.to_excel(path, index=False, sheet_name="Ventas")
    print(f"  Created: {path.name}")


def generate_excel_dist_b(sales_df: pd.DataFrame) -> None:
    """
    Distributor B — MediFar: Pivot-style with merged cells header (complex).
    Month-level summary table with product columns.
    """
    mask = sales_df["distributor_code"] == "DIST_B"
    sample = sales_df[mask].head(300).copy()

    with pd.ExcelWriter(EXCEL_DIR / "DIST_B_MediFar_reporte.xlsx", engine="openpyxl") as writer:
        # Sheet 1: Summary pivot
        pivot = sample.groupby(["full_date", "client_name", "product_name"])["net_amount"].sum().unstack(fill_value=0)
        pivot.reset_index(inplace=True)
        pivot.to_excel(writer, sheet_name="Resumen_Pivot", index=False)

        # Sheet 2: Detail
        detail = pd.DataFrame({
            "FECHA":       sample["full_date"],
            "FACTURA":     sample["invoice_number"],
            "COD_MF":      "MF" + sample["product_id"].str.replace("SAP-", "").str.zfill(4),
            "PRODUCTO":    sample["product_name"],
            "CLIENTE":     sample["client_name"],
            "CANT":        sample["quantity"],
            "PRECIO":      sample["unit_price"],
            "IMPORTE":     sample["net_amount"],
            "PROV":        sample["city_name"],
        })
        detail.to_excel(writer, sheet_name="Detalle", index=False)

    print(f"  Created: DIST_B_MediFar_reporte.xlsx")


def generate_excel_dist_c(sales_df: pd.DataFrame) -> None:
    """
    Distributor C — Farmacorp: CSV saved as .xlsx, semicolon-separated inside.
    Headers in row 3 (first 2 rows are company header).
    """
    from openpyxl import Workbook
    mask = sales_df["distributor_code"] == "DIST_C"
    sample = sales_df[mask].head(200).copy()

    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    # Company header rows
    ws["A1"] = "FARMACORP S.A."
    ws["A2"] = f"Reporte de Ventas Exportado: {datetime.now().strftime('%d/%m/%Y')}"
    ws["A3"] = ""

    # Actual headers in row 4
    headers = ["fecha","numero_doc","codigo_fc","descripcion_prod","qty","precio_venta",
               "total","nombre_cliente","ciudad_cliente","descuento"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)

    # Data from row 5
    for row_idx, (_, row) in enumerate(sample.iterrows(), 5):
        ws.cell(row=row_idx, column=1, value=row["full_date"])
        ws.cell(row=row_idx, column=2, value=row["invoice_number"])
        ws.cell(row=row_idx, column=3, value=f"FC-{row['product_id'].replace('SAP-','')}")
        ws.cell(row=row_idx, column=4, value=row["product_name"])
        ws.cell(row=row_idx, column=5, value=row["quantity"])
        ws.cell(row=row_idx, column=6, value=row["unit_price"])
        ws.cell(row=row_idx, column=7, value=row["net_amount"])
        ws.cell(row=row_idx, column=8, value=row["client_name"])
        ws.cell(row=row_idx, column=9, value=row["city_name"])
        ws.cell(row=row_idx, column=10, value=f"{row['discount_pct']}%")

    wb.save(EXCEL_DIR / "DIST_C_Farmacorp_export.xlsx")
    print(f"  Created: DIST_C_Farmacorp_export.xlsx")


def generate_excel_dist_d(sales_df: pd.DataFrame) -> None:
    """
    Distributor D — AlphaFarma: Multi-sheet workbook, one sheet per zone.
    """
    mask = sales_df["distributor_code"] == "DIST_D"
    sample = sales_df[mask].head(400).copy()

    with pd.ExcelWriter(EXCEL_DIR / "DIST_D_AlphaFarma_multizona.xlsx", engine="openpyxl") as writer:
        for zone in ["CAP", "NOR", "EST", "SUR", "OES"]:
            zone_data = sample[sample["zone_code"] == zone]
            if zone_data.empty:
                continue
            df_zone = pd.DataFrame({
                "Fecha":       zone_data["full_date"],
                "Factura":     zone_data["invoice_number"],
                "Codigo":      "ALF" + zone_data["product_id"].str.replace("SAP-", "").str.zfill(3) + "A",
                "Producto":    zone_data["product_name"],
                "Categoria":   zone_data["category"],
                "Cliente":     zone_data["client_name"],
                "TipoCliente": zone_data["client_type"],
                "Ciudad":      zone_data["city_name"],
                "Cantidad":    zone_data["quantity"],
                "PVP":         zone_data["unit_price"],
                "Neto":        zone_data["net_amount"],
            })
            df_zone.to_excel(writer, sheet_name=f"Zona_{zone}", index=False)

    print(f"  Created: DIST_D_AlphaFarma_multizona.xlsx")


def generate_excel_dist_e(sales_df: pd.DataFrame) -> None:
    """
    Distributor E — BioPharma: Dynamic header row (header not always row 1).
    Contains metadata rows interspersed. Header at row 6.
    """
    from openpyxl import Workbook
    mask = sales_df["distributor_code"] == "DIST_E"
    sample = sales_df[mask].head(200).copy()

    wb = Workbook()
    ws = wb.active
    ws.title = "ReporteVentas"

    # Metadata block
    ws["A1"] = "BIOPHARMA DISTRIBUTION S.A."
    ws["A2"] = "RNC: 1-01-67890-1"
    ws["A3"] = f"Período: {sample['full_date'].min()} al {sample['full_date'].max()}"
    ws["A4"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A5"] = "DETALLE DE VENTAS"

    # Headers at row 6
    hdrs = ["Fecha_Venta","ID_Factura","CodBP","Nombre_Medicamento","Unidades",
            "Precio","Descto_Pct","Importe_Neto","Razon_Social","Municipio",
            "Lab","Categoria_Terapeutica"]
    for c, h in enumerate(hdrs, 1):
        ws.cell(row=6, column=c, value=h)

    for r, (_, row) in enumerate(sample.iterrows(), 7):
        vals = [
            row["full_date"], row["invoice_number"],
            f"BP{row['product_id'].replace('SAP-','').zfill(2)}X",
            row["product_name"], row["quantity"], row["unit_price"],
            row["discount_pct"], row["net_amount"],
            row["client_name"], row["city_name"],
            row["lab_code"], row["category"],
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v)

    wb.save(EXCEL_DIR / "DIST_E_BioPharma_dinamico.xlsx")
    print(f"  Created: DIST_E_BioPharma_dinamico.xlsx")


def generate_excel_dist_f(sales_df: pd.DataFrame) -> None:
    """
    Distributor F — MedDist: City and zone combined in one column "Ciudad-Zona".
    Also uses non-standard date format DD/MM/YYYY and comma decimal separators.
    """
    mask = sales_df["distributor_code"] == "DIST_F"
    sample = sales_df[mask].head(200).copy()

    df = pd.DataFrame({
        "Fecha":           pd.to_datetime(sample["full_date"]).dt.strftime("%d/%m/%Y"),
        "Referencia":      sample["invoice_number"],
        "CodMD":           "MD-" + sample["product_id"].str.replace("SAP-", "").str.zfill(5),
        "Medicamento":     sample["product_name"],
        "ClienteNombre":   sample["client_name"],
        "Ciudad_Zona":     sample["city_name"] + " / " + sample["zone_code"],
        "Unidades":        sample["quantity"],
        # Comma decimal separator
        "Precio_Unit":     sample["unit_price"].apply(lambda x: str(x).replace(".", ",")),
        "Importe":         sample["net_amount"].apply(lambda x: str(x).replace(".", ",")),
        "Margen_Aprox":    sample["margin_pct"].apply(lambda x: f"{x:.2f}%".replace(".", ",")),
        "Vendedor_Ext":    "MedDist Equipo " + sample["zone_code"],
    })

    path = EXCEL_DIR / "DIST_F_MedDist_combinado.xlsx"
    df.to_excel(path, index=False, sheet_name="Ventas_MedDist")
    print(f"  Created: {path.name}")


# ════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("Pharma DR · Synthetic Data Generator")
    print("=" * 60)

    print("\n[1/5] Generating 200 clients...")
    clients = generate_clients(200)
    clients.to_csv(OUTPUT_DIR / "clients.csv", index=False, encoding="utf-8-sig")
    print(f"      → {len(clients)} clients written to clients.csv")

    print("\n[2/5] Generating ~100k sales rows (2021–2024)...")
    sales = generate_sales(clients, n_invoices=25000)
    sales.to_csv(OUTPUT_DIR / "sales.csv", index=False, encoding="utf-8-sig")
    print(f"      → {len(sales):,} rows written to sales.csv")
    print(f"      → Total net sales: RD$ {sales['net_amount'].sum():,.0f}")
    print(f"      → Date range: {sales['full_date'].min()} → {sales['full_date'].max()}")

    print("\n[3/5] Generating monthly targets (2021–2024)...")
    all_targets = []
    for year in range(2021, 2025):
        for month in range(1, 13):
            all_targets.extend(generate_targets(year, month))
    targets_df = pd.DataFrame(all_targets)
    targets_df.to_csv(OUTPUT_DIR / "targets.csv", index=False, encoding="utf-8-sig")
    print(f"      → {len(targets_df)} target rows written to targets.csv")

    print("\n[4/5] Generating product mapping table...")
    mapping = generate_product_mapping()
    mapping.to_csv(MAPPING_DIR / "product_mapping.csv", index=False, encoding="utf-8-sig")
    print(f"      → {len(mapping)} mapping rows written to product_mapping.csv")

    print("\n[5/5] Generating 6 Excel distributor format samples...")
    generate_excel_dist_a(sales)
    generate_excel_dist_b(sales)
    generate_excel_dist_c(sales)
    generate_excel_dist_d(sales)
    generate_excel_dist_e(sales)
    generate_excel_dist_f(sales)

    print("\n" + "=" * 60)
    print("Data generation complete!")
    print(f"Outputs in: {OUTPUT_DIR}")
    print(f"Excel samples in: {EXCEL_DIR}")
    print(f"Product mapping in: {MAPPING_DIR}")
    print("=" * 60)

    # Summary stats
    print("\nSales Summary by Year:")
    yearly = sales.groupby(
        pd.to_datetime(sales["full_date"]).dt.year
    )["net_amount"].agg(["sum", "count"])
    print(yearly.rename(columns={"sum": "Total Net (RD$)", "count": "Rows"}))

    print("\nSales by Zone:")
    zone_sum = sales.groupby("zone_code")["net_amount"].sum().sort_values(ascending=False)
    for z, v in zone_sum.items():
        print(f"  {z}: RD$ {v:>15,.0f}")

    print("\nTop 10 Products by Net Sales:")
    top_prod = sales.groupby("product_name")["net_amount"].sum().sort_values(ascending=False).head(10)
    for p, v in top_prod.items():
        print(f"  {p[:40]:<40} RD$ {v:>12,.0f}")


if __name__ == "__main__":
    main()
